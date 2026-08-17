"""Read the contact spreadsheet the client actually works from.

His sheet is not a flat table. It is grouped for reading:

    Vilaow — Professional Outreach List          <- title
    Real, publicly listed businesses ...         <- description
    CRETE (Chania · Heraklion · ...)             <- region band
    Chania  (67 contacts)                        <- city band
    Category | Business / Name | Phone | ...     <- header, REPEATED per city
    Estate Agent | ARENCORES ... | +30 ...       <- data
    ...
    Heraklion  (65 contacts)                     <- next city band
    Category | ...                               <- header again

A parser that assumes "row 1 is the header, the rest is data" ingests the
11 band rows and 6 repeated headers as if they were 17 people called things
like "Chania  (67 contacts)". It also loses the city entirely, because there
is no city column — the city is the band you are standing in.

So this walks the rows keeping track of the current region and city, and only
accepts a row whose first cell is a known profession.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field

from openpyxl import load_workbook

# The six he uses, plus the spellings his sheet actually contains. "Accountant"
# in his sheet is our "tax accountant" — the only category whose name differs.
CATEGORY_TO_KEY = {
    "estate agent": "agent",
    "lawyer": "lawyer",
    "civil engineer": "engineer",
    "architect": "architect",
    "contractor": "contractor",
    "accountant": "accountant",
    "tax accountant": "accountant",
}

# A region band is written in capitals; a city band is title case with a count.
REGION_BAND = re.compile(r"^[A-Z][A-Z\s·]+$")


@dataclass
class Row:
    profession_key: str
    business_name: str
    phone: str | None
    address: str | None
    city: str | None
    region: str | None
    rating: float | None
    review_count: int | None
    status: str | None
    notes: str | None


@dataclass
class ImportResult:
    rows: list[Row] = field(default_factory=list)
    seen: int = 0
    skipped_no_phone: int = 0
    skipped_unknown_category: int = 0
    band_rows: int = 0
    header_rows: int = 0

    @property
    def accepted(self) -> int:
        return len(self.rows)


def _text(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _number(v, cast):
    try:
        return cast(str(v).strip())
    except (TypeError, ValueError):
        return None


def parse_workbook(path_or_stream, sheet_name: str = "Contacts") -> ImportResult:
    wb = load_workbook(path_or_stream, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]

    out = ImportResult()
    region: str | None = None
    city: str | None = None

    for raw in ws.iter_rows(values_only=True):
        cells = list(raw) + [None] * (8 - len(raw))
        first = _text(cells[0])
        if not first:
            continue

        # The header, repeated once per city block.
        if first.lower() == "category":
            out.header_rows += 1
            continue

        # A band row: only the first cell carries anything. Three kinds, and
        # telling them apart matters — an earlier version treated anything
        # non-capitalised as a city, which made the sheet's title row a city
        # and, because "CRETE (Chania · ...)" contains a bracket, left every
        # Cretan row with no region at all: 216 of 323.
        if all(_text(c) is None for c in cells[1:]):
            out.band_rows += 1
            head = first.split("(")[0].strip()
            if head.isupper():
                region = head.title()                    # CRETE -> Crete
            elif "contact" in first.lower() and "(" in first:
                city = head or None                      # Chania  (67 contacts)
            # anything else is the sheet's title or description — ignore it
            continue

        out.seen += 1
        key = CATEGORY_TO_KEY.get(first.lower())
        if key is None:
            out.skipped_unknown_category += 1
            continue

        phone = _text(cells[2])
        if not phone:
            # He notes these himself: Google does not list a number publicly
            # for some businesses. Without one there is nothing to call.
            out.skipped_no_phone += 1
            continue

        out.rows.append(
            Row(
                profession_key=key,
                business_name=_text(cells[1]) or "(no name)",
                phone=phone,
                address=_text(cells[3]),
                city=city,
                region=region,
                rating=_number(cells[4], float),
                review_count=_number(cells[5], int),
                status=_text(cells[6]),
                notes=_text(cells[7]),
            )
        )

    wb.close()
    return out


def normalise_phone(phone: str) -> str:
    """For duplicate detection only — the original is what gets dialled.

    His sheet mixes "+30 697 044 7994" and "6947713390"; the same business
    imported twice must not become two rows to ring.
    """
    digits = re.sub(r"\D", "", phone or "")
    if digits.startswith("0030"):
        digits = digits[4:]
    elif digits.startswith("30") and len(digits) > 10:
        digits = digits[2:]
    return digits
